import logging
import os
import tempfile
from datetime import datetime
from io import BytesIO
from typing import List, Optional, Tuple

import openpyxl
import pandas as pd
import requests
from flask import Flask, jsonify, make_response, request
from joblib import dump, load
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.linear_model import LogisticRegression
from sklearn.model_selection import train_test_split
from sklearn.pipeline import Pipeline
from werkzeug.exceptions import HTTPException

app = Flask(__name__)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    handlers=[logging.FileHandler("name_classifier.log"), logging.StreamHandler()],
)
logger = logging.getLogger(__name__)

# Constants
MODEL_FILE = "name_classifier.joblib"
TRAINING_DATA_URLS = [
    "https://raw.githubusercontent.com/ErnestoAizenberg/kz-ru-name-detector/refs/heads/main/ru_names_1.xlsx",
    "https://raw.githubusercontent.com/ErnestoAizenberg/kz-ru-name-detector/refs/heads/main/kz_names_1.xlsx",
    "https://raw.githubusercontent.com/ErnestoAizenberg/kz-ru-name-detector/refs/heads/main/kz_names_2.xlsx",
]


class ModelTrainingError(Exception):
    """Custom exception for model training failures"""

    pass


class FileProcessingError(Exception):
    """Custom exception for file processing failures"""

    pass


def setup_model() -> Pipeline:
    """Initialize or load the classification model"""
    try:
        if not os.path.exists(MODEL_FILE):
            logger.info("Model not found. Training new model...")
            model = train_model()
            dump(model, MODEL_FILE)
            logger.info(f"Model successfully trained and saved to {MODEL_FILE}")
        else:
            logger.info(f"Loading model from {MODEL_FILE}")
            model = load(MODEL_FILE)
        return model
    except Exception as e:
        logger.error(f"Failed to setup model: {str(e)}")
        raise ModelTrainingError("Could not initialize the classification model")


def download_training_data() -> None:
    """Download training data files"""
    try:
        for url in TRAINING_DATA_URLS:
            filename = url.split("/")[-1]
            logger.info(f"Downloading training data: {filename}")
            response = requests.get(url, timeout=10)
            response.raise_for_status()

            with open(filename, "wb") as f:
                f.write(response.content)
            logger.info(f"Successfully downloaded {filename}")
    except requests.exceptions.RequestException as e:
        logger.error(f"Failed to download training data: {str(e)}")
        raise ModelTrainingError("Could not download training data")


def df_to_string_list(df: pd.DataFrame) -> List[str]:
    """Convert DataFrame rows to list of strings"""
    try:
        return [
            " ".join(str(cell).strip() if pd.notna(cell) else "" for cell in row)
            for row in df.values.tolist()
        ]
    except Exception as e:
        logger.error(f"DataFrame conversion failed: {str(e)}")
        raise ModelTrainingError("Could not process training data")


def train_model() -> Pipeline:
    """Train and return a new classification model"""
    try:
        download_training_data()

        # Load and process training data
        df_ru = pd.read_excel("ru_names_1.xlsx", usecols=[0, 1, 2], header=None)
        df_kz1 = pd.read_excel("kz_names_1.xlsx", usecols=[0, 1, 2], header=None)
        df_kz2 = pd.read_excel("kz_names_2.xlsx", usecols=[3], header=None)

        list_ru = df_to_string_list(df_ru)
        list_kz = df_to_string_list(df_kz1) + df_to_string_list(df_kz2)

        # Create training DataFrame
        df = pd.concat(
            [
                pd.DataFrame({"name": list_kz, "label": "kz"}),
                pd.DataFrame({"name": list_ru, "label": "ru"}),
            ]
        )

        # Train model
        model = Pipeline(
            [
                ("tfidf", TfidfVectorizer(analyzer="char", ngram_range=(1, 3))),
                ("clf", LogisticRegression(max_iter=1000)),
            ]
        )

        X_train, X_test, y_train, y_test = train_test_split(
            df["name"], df["label"], test_size=0.2, random_state=42
        )

        model.fit(X_train, y_train)
        accuracy = model.score(X_test, y_test)
        logger.info(f"Model trained with accuracy: {accuracy:.2f}")

        return model
    except Exception as e:
        logger.error(f"Model training failed: {str(e)}")
        raise ModelTrainingError("Model training process failed")


def has_kazakh_letters(name: str) -> bool:
    """Check if name contains Kazakh-specific letters"""
    kazakh_letters = {
        "Ә",
        "ә",
        "Ғ",
        "ғ",
        "Қ",
        "қ",
        "Ң",
        "ң",
        "Ө",
        "ө",
        "Ұ",
        "ұ",
        "Ү",
        "ү",
        "Һ",
        "һ",
        "І",
        "і",
    }
    return any(char in kazakh_letters for char in name)


def process_uploaded_file(
    file_stream,
    name_columns: List[int],
    country_filter: Optional[str] = None,
    address_columns: Optional[List[int]] = None,
    search_value: Optional[str] = None,
) -> Tuple[BytesIO, int]:
    """
    Process the uploaded Excel file with enhanced filtering options

    Args:
        file_stream: Uploaded file stream
        name_columns: List of column indices containing name parts
        country_filter: 'kz' or 'ru' to filter by country
        address_columns: List of column indices containing address information
        search_value: Value to search in address columns

    Returns:
        Tuple containing: (output file buffer, count of matching records)
    """
    try:
        with tempfile.NamedTemporaryFile(suffix=".xlsx") as tmp_source:
            # Save uploaded file
            file_stream.save(tmp_source.name)

            # Process the file
            try:
                wb_source = openpyxl.load_workbook(tmp_source.name)
                ws_source = wb_source.active
            except Exception as e:
                logger.error(f"Invalid Excel file: {str(e)}")
                raise FileProcessingError("The uploaded file is not a valid Excel file")

            # Prepare output file
            wb_output = openpyxl.Workbook()
            ws_output = wb_output.active

            # Copy headers if they exist
            if ws_source.max_row > 0:
                headers = [cell.value for cell in ws_source[1]]
                ws_output.append(headers)

            # Process rows
            match_count = 0
            processed_rows = 0

            for row in ws_source.iter_rows(min_row=2, values_only=True):
                processed_rows += 1
                try:
                    # Check address search if provided
                    if search_value and address_columns:
                        address_parts = [
                            str(row[col]).lower()
                            for col in address_columns
                            if col < len(row)
                        ]
                        full_address = " ".join(address_parts)
                        if search_value.lower() not in full_address:
                            continue

                    # Process name classification if country filter is set
                    if country_filter:
                        full_name = " ".join(
                            str(row[col]).strip()
                            for col in name_columns
                            if col < len(row) and row[col] is not None
                        ).strip()

                        if not full_name:
                            continue

                        # First check for Kazakh letters if we're looking for KZ names
                        if country_filter == "kz" and has_kazakh_letters(full_name):
                            ws_output.append(row)
                            match_count += 1
                            continue

                        # Use model for prediction
                        prediction = model.predict([full_name])[0]
                        if prediction != country_filter:
                            continue

                    # If we get here, the row matches all criteria
                    ws_output.append(row)
                    match_count += 1

                except Exception as e:
                    logger.warning(f"Error processing row {processed_rows}: {str(e)}")
                    continue

            logger.info(
                f"Processed {processed_rows} rows. Matching records: {match_count}"
            )

            # Save results to buffer
            output_buffer = BytesIO()
            wb_output.save(output_buffer)
            output_buffer.seek(0)

            return output_buffer, match_count

    except Exception as e:
        logger.error(f"File processing failed: {str(e)}")
        raise FileProcessingError("Could not process the uploaded file")


# Error handlers (remain the same as before)
@app.errorhandler(HTTPException)
def handle_http_error(e):
    logger.warning(f"HTTP error {e.code}: {e.description}")
    return jsonify({"error": e.name, "message": e.description}), e.code


@app.errorhandler(ModelTrainingError)
def handle_model_error(e):
    logger.error(f"Model error: {str(e)}")
    return jsonify({"error": "Model Error", "message": str(e)}), 500


@app.errorhandler(FileProcessingError)
def handle_file_error(e):
    logger.error(f"File processing error: {str(e)}")
    return jsonify({"error": "File Processing Error", "message": str(e)}), 400


@app.errorhandler(Exception)
def handle_unexpected_error(e):
    logger.critical(f"Unexpected error: {str(e)}", exc_info=True)
    return (
        jsonify(
            {
                "error": "Internal Server Error",
                "message": "An unexpected error occurred",
            }
        ),
        500,
    )


# Initialize model
try:
    model = setup_model()
except ModelTrainingError:
    logger.critical("Application failed to start due to model initialization error")
    raise


@app.route("/process_names", methods=["POST"])
def process_names():
    """API endpoint for processing name files with enhanced filtering"""
    try:
        logger.info("Name processing request received")

        # Validate request
        if "file" not in request.files:
            logger.warning("No file uploaded")
            raise FileProcessingError("No file uploaded")

        file = request.files["file"]
        if file.filename == "":
            logger.warning("Empty filename")
            raise FileProcessingError("No file selected")

        if not file.filename.lower().endswith((".xlsx", ".xls")):
            logger.warning(f"Invalid file type: {file.filename}")
            raise FileProcessingError("Only Excel files (.xlsx, .xls) are supported")

        # Get parameters
        name_columns_str = request.form.get("name_columns", "0,1,2")
        country_filter = request.form.get("country_filter")  # 'kz', 'ru', or None
        address_columns_str = request.form.get("address_columns", "")
        search_value = request.form.get("search_value", "").strip()

        try:
            name_columns = [int(col.strip()) for col in name_columns_str.split(",")]
            address_columns = (
                [int(col.strip()) for col in address_columns_str.split(",")]
                if address_columns_str
                else None
            )
        except ValueError as e:
            raise FileProcessingError(f"Invalid column numbers: {str(e)}")

        # Validate country filter
        if country_filter and country_filter not in ("kz", "ru"):
            raise FileProcessingError(
                "Invalid country filter. Use 'kz' or 'ru' or leave empty"
            )

        # Validate search parameters
        if search_value and not address_columns:
            raise FileProcessingError(
                "Address columns must be specified when using address search"
            )

        # Process file
        try:
            output_buffer, match_count = process_uploaded_file(
                file,
                name_columns,
                country_filter,
                address_columns,
                search_value if search_value else None,
            )
        except Exception as e:
            logger.error(f"File processing failed: {str(e)}")
            raise FileProcessingError("Could not process the file")

        # Prepare response
        response = make_response(output_buffer.getvalue())
        response.headers.set(
            "Content-Type",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        filename_parts = ["filtered_names"]
        if country_filter:
            filename_parts.append(country_filter)
        if search_value:
            filename_parts.append("search")
        filename_parts.append(datetime.now().strftime("%Y%m%d_%H%M%S"))

        response.headers.set(
            "Content-Disposition",
            "attachment",
            filename=f"{'_'.join(filename_parts)}.xlsx",
        )

        logger.info(f"Successfully processed file with {match_count} matching records")
        return response

    except FileProcessingError:
        raise
    except Exception as e:
        logger.error(f"Unexpected processing error: {str(e)}")
        raise FileProcessingError("An unexpected error occurred during processing")


@app.route("/classify_name", methods=["POST"])
def classify_name():
    """Endpoint for classifying a single name"""
    try:
        data = request.get_json()
        name = data.get("name", "").strip()

        if not name:
            return jsonify({"error": "Name is required"}), 400

        # First check for Kazakh letters
        if has_kazakh_letters(name):
            return jsonify(
                {
                    "name": name,
                    "classification": "kz",
                    "method": "kazakh_letters_check",
                    "probability": 1.0,
                }
            )

        # Use model for prediction
        prediction = model.predict([name])[0]
        proba = model.predict_proba([name])[0]
        probability = proba[0] if prediction == "kz" else proba[1]

        return jsonify(
            {
                "name": name,
                "classification": prediction,
                "method": "model_prediction",
                "probability": float(probability),
            }
        )

    except Exception as e:
        logger.error(f"Error classifying name: {str(e)}")
        return jsonify({"error": "Internal server error"}), 500


@app.route("/")
def index():
    """Main page with enhanced upload form including address search"""
    logger.info("Serving index page")
    return """
    <!DOCTYPE html>
    <html lang="en">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Name Classifier (KZ/RU) with Address Search</title>
        <style>
            :root {
                --primary: #4361ee;
                --primary-hover: #3a56d4;
                --secondary: #f72585;
                --light: #f8f9fa;
                --dark: #212529;
                --gray: #6c757d;
                --success: #4cc9f0;
                --warning: #f8961e;
                --border-radius: 12px;
                --box-shadow: 0 4px 20px rgba(0, 0, 0, 0.08);
                --transition: all 0.3s cubic-bezier(0.25, 0.8, 0.25, 1);
            }

            body {
                font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif;
                line-height: 1.5;
                max-width: 800px;
                margin: 0 auto;
                padding: 2rem;
                color: var(--dark);
                background-color: #f5f7fa;
                -webkit-font-smoothing: antialiased;
            }

            h1 {
                color: var(--dark);
                text-align: center;
                margin-bottom: 2rem;
                font-weight: 700;
                font-size: 2.2rem;
                letter-spacing: -0.5px;
            }

            .form-container {
                background: white;
                padding: 2rem;
                border-radius: var(--border-radius);
                box-shadow: var(--box-shadow);
                margin-bottom: 2rem;
                transition: var(--transition);
                border: 1px solid rgba(0, 0, 0, 0.03);
            }

            .form-container:hover {
                box-shadow: 0 6px 24px rgba(0, 0, 0, 0.1);
            }

            .form-group {
                margin-bottom: 1.5rem;
            }

            label {
                display: block;
                margin-bottom: 0.5rem;
                font-weight: 600;
                color: var(--dark);
                font-size: 0.95rem;
            }

            input[type="file"],
            input[type="text"],
            select {
                width: 100%;
                padding: 0.75rem 1rem;
                border: 1px solid #e0e0e0;
                border-radius: var(--border-radius);
                box-sizing: border-box;
                font-family: inherit;
                font-size: 1rem;
                transition: var(--transition);
                background-color: var(--light);
            }

            input[type="file"]:focus,
            input[type="text"]:focus,
            select:focus {
                outline: none;
                border-color: var(--primary);
                box-shadow: 0 0 0 3px rgba(67, 97, 238, 0.2);
            }

            input[type="submit"],
            .submit-btn {
                background: var(--primary);
                color: white;
                border: none;
                padding: 0.75rem 2rem;
                border-radius: var(--border-radius);
                cursor: pointer;
                font-size: 1rem;
                font-weight: 600;
                display: inline-block;
                margin: 1.5rem auto 0;
                width: auto;
                transition: var(--transition);
                box-shadow: 0 2px 10px rgba(67, 97, 238, 0.3);
            }

            input[type="submit"]:hover,
            .submit-btn:hover {
                background: var(--primary-hover);
                box-shadow: 0 4px 14px rgba(67, 97, 238, 0.4);
                transform: translateY(-1px);
            }

            .filter-section {
                background: rgba(67, 97, 238, 0.05);
                padding: 1.5rem;
                border-radius: var(--border-radius);
                margin: 1.5rem 0;
                border: 1px solid rgba(67, 97, 238, 0.1);
            }

            .filter-section h3 {
                margin-top: 0;
                color: var(--primary);
                font-size: 1.1rem;
                font-weight: 600;
                margin-bottom: 1rem;
            }

            .file-info, small {
                color: var(--gray);
                font-size: 0.85rem;
                display: block;
                margin-top: 0.5rem;
            }

            .result-kz {
                background-color: rgba(76, 201, 240, 0.1);
                border: 1px solid var(--success);
                border-left: 4px solid var(--success);
            }

            .result-ru {
                background-color: rgba(248, 150, 30, 0.1);
                border: 1px solid var(--warning);
                border-left: 4px solid var(--warning);
            }

            .probability {
                margin-top: 1rem;
                height: 8px;
                background: #e9ecef;
                border-radius: 4px;
                overflow: hidden;
            }

            .probability-bar {
                height: 100%;
                background: linear-gradient(90deg, var(--primary), var(--secondary));
                width: 0%;
                transition: width 0.7s ease-out;
            }

            /* Анимации */
            @keyframes fadeIn {
                from { opacity: 0; transform: translateY(10px); }
                to { opacity: 1; transform: translateY(0); }
            }

            .form-container {
                animation: fadeIn 0.6s ease-out;
            }

            /* Кастомизация файлового инпута */
            .file-input-wrapper {
                position: relative;
                overflow: hidden;
                display: inline-block;
                width: 100%;
            }

            .file-input-button {
                background: var(--light);
                border: 1px dashed #ced4da;
                border-radius: var(--border-radius);
                padding: 2rem;
                text-align: center;
                cursor: pointer;
                transition: var(--transition);
            }

            .file-input-button:hover {
                background: #e9ecef;
                border-color: var(--primary);
            }

            .file-input-wrapper input[type="file"] {
                position: absolute;
                left: 0;
                top: 0;
                opacity: 0;
                width: 100%;
                height: 100%;
                cursor: pointer;
            }

            @media (max-width: 768px) {
                body {
                    padding: 1rem;
                }

                h1 {
                    font-size: 1.8rem;
                }

                .form-container {
                    padding: 1.5rem;
                }
            }
        </style>
    </head>
    <body>
        <h1>Name Classifier (KZ/RU) with Address Search</h1>

        <div class="form-container">
            <form action="/process_names" method="post" enctype="multipart/form-data" id="uploadForm">
                <div class="form-group">
                    <label for="file">Upload Excel file with names:</label>
                    <input type="file" id="file" name="file" accept=".xlsx,.xls" required>
                    <div class="file-info">Accepted formats: .xlsx, .xls</div>
                </div>

                <div class="form-group">
                    <label for="name_columns">Columns containing name parts (0-based, comma-separated):</label>
                    <input type="text" id="name_columns" name="name_columns" value="0,1,2" required>
                    <small>Example: For first name in column 0 and last name in column 1, use "0,1"</small>
                </div>

                <div class="filter-section">
                    <h3>Name Filter</h3>
                    <div class="form-group">
                        <label for="country_filter">Filter by country:</label>
                        <select id="country_filter" name="country_filter">
                            <option value="">All countries</option>
                            <option value="kz">Kazakh (KZ) only</option>
                            <option value="ru">Russian (RU) only</option>
                        </select>
                    </div>
                </div>

                <div class="filter-section">
                    <h3>Address Search</h3>
                    <div class="form-group">
                        <label for="address_columns">Columns containing address (0-based, comma-separated):</label>
                        <input type="text" id="address_columns" name="address_columns">
                        <small>Leave empty to disable address search</small>
                    </div>
                    <div class="form-group">
                        <label for="search_value">Search value in addresses:</label>
                        <input type="text" id="search_value" name="search_value">
                        <small>Enter text to search in address columns</small>
                    </div>
                </div>

                <input type="submit" value="Process" id="submitBtn">
            </form>
        </div>

        <div class="form-container">
            <h3 style="text-align: center; margin-bottom: 15px;">Test Single Name</h3>
            <form id="singleNameForm" onsubmit="checkSingleName(event)">
                <div class="form-group">
                    <label for="test_name">Enter a name to classify:</label>
                    <input type="text" id="test_name" name="test_name" required>
                </div>
                <input type="submit" value="Check" class="submit-btn">
            </form>
            <div id="result" style="margin-top: 20px; padding: 15px; border-radius: 5px; display: none;"></div>
        </div>

        <script>
            document.getElementById('uploadForm').addEventListener('submit', function(e) {
                const fileInput = document.getElementById('file');
                const columnsInput = document.getElementById('name_columns');
                const addressColumnsInput = document.getElementById('address_columns');
                const searchValueInput = document.getElementById('search_value');
                const submitBtn = document.getElementById('submitBtn');

                // Validate file extension
                const fileName = fileInput.value.toLowerCase();
                if (!fileName.endsWith('.xlsx') && !fileName.endsWith('.xls')) {
                    alert('Please upload an Excel file (.xlsx or .xls)');
                    e.preventDefault();
                    return;
                }

                // Validate name columns format
                if (!/^\d+(,\d+)*$/.test(columnsInput.value)) {
                    alert('Please enter valid name column numbers (e.g. "0,1,2")');
                    e.preventDefault();
                    return;
                }

                // Validate address search parameters
                if (searchValueInput.value.trim() && !addressColumnsInput.value.trim()) {
                    alert('Please specify address columns when using address search');
                    e.preventDefault();
                    return;
                }

                if (addressColumnsInput.value.trim() && !/^\d+(,\d+)*$/.test(addressColumnsInput.value)) {
                    alert('Please enter valid address column numbers (e.g. "3,4")');
                    e.preventDefault();
                    return;
                }

                // Disable button to prevent double submission
                submitBtn.disabled = true;
                submitBtn.value = 'Processing...';
            });

            function checkSingleName(event) {
                event.preventDefault();
                const name = document.getElementById('test_name').value.trim();
                if (!name) return;

                fetch('/classify_name', {
                    method: 'POST',
                    headers: {
                        'Content-Type': 'application/json',
                    },
                    body: JSON.stringify({name: name})
                })
                .then(response => response.json())
                .then(data => {
                    const resultDiv = document.getElementById('result');
                    resultDiv.style.display = 'block';

                    if (data.classification === 'kz') {
                        resultDiv.className = 'result result-kz';
                        resultDiv.innerHTML = `
                            <h4>Result for: ${data.name}</h4>
                            <p><strong>Classification:</strong> Kazakh (KZ)</p>
                            ${data.method === 'kazakh_letters_check' ?
                              '<p>Detected Kazakh letters</p>' :
                              `<p>Model prediction (${(data.probability * 100).toFixed(1)}% confidence)</p>`
                            }
                        `;
                    } else {
                        resultDiv.className = 'result result-ru';
                        resultDiv.innerHTML = `
                            <h4>Result for: ${data.name}</h4>
                            <p><strong>Classification:</strong> Russian (RU)</p>
                            <p>Model prediction (${(data.probability * 100).toFixed(1)}% confidence)</p>
                            <div class="probability">
                                <div class="probability-bar" style="width: ${data.probability * 100}%"></div>
                            </div>
                        `;
                    }
                })
                .catch(error => {
                    console.error('Error:', error);
                    alert('An error occurred while processing the name');
                });
            }
        </script>
    </body>
    </html>
    """


if __name__ == "__main__":
    try:
        app.run(host="0.0.0.0", port=5000)
    except Exception as e:
        logger.critical(f"Application failed to start: {str(e)}")
        raise
