from typing import List

import openpyxl
import pandas as pd
import requests
from joblib import dump
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.linear_model import LogisticRegression
from sklearn.model_selection import train_test_split
from sklearn.pipeline import Pipeline

urls = [
    "https://raw.githubusercontent.com/ErnestoAizenberg/kz-ru-name-detector/refs/heads/main/ru_names_1.xlsx",
    "https://raw.githubusercontent.com/ErnestoAizenberg/kz-ru-name-detector/refs/heads/main/kz_names_1.xlsx",
    "https://raw.githubusercontent.com/ErnestoAizenberg/kz-ru-name-detector/refs/heads/main/kz_names_2.xlsx",
]


# Функция для преобразования в список строк
def df_to_string_list(df):
    return [
        " ".join(str(cell).strip() if pd.notna(cell) else "" for cell in row)
        for row in df.values.tolist()
    ]


def join_cols_to_full_name(row: tuple, usecols: List[int]) -> str:
    """Combine specified columns from a row into a full name string.

    Args:
        row: A tuple containing all cell values from a worksheet row
        usecols: List of column indices to include in the full name

    Returns:
        A string containing the combined values from specified columns
    """
    return " ".join(str(row[col]) for col in usecols).strip()


def classify_and_split_names(
    model, source_file: str, name_columns: List[int], preview_rows: int = 5
) -> None:
    """Classify names in an Excel file and split them into separate files by nationality.

    Args:
        model: Class with predict method
        source_file: Path to the source Excel file
        name_columns: List of column indices that contain name parts to be combined
        preview_rows: Number of rows to preview from each output file (0 to disable)
    """
    wb_source = openpyxl.load_workbook(source_file)
    ws_source = wb_source.active

    # Create two new files for writing
    wb_target_kz = openpyxl.Workbook()
    ws_target_kz = wb_target_kz.active
    wb_target_ru = openpyxl.Workbook()
    ws_target_ru = wb_target_ru.active

    # Copy headers if they exist
    headers = [cell.value for cell in ws_source[1]]
    ws_target_kz.append(headers)
    ws_target_ru.append(headers)

    # Iterate through rows of the source file
    for row in ws_source.iter_rows(min_row=2, values_only=True):
        person_full_name = join_cols_to_full_name(row, name_columns)
        pred = model.predict([person_full_name])
        if pred == "kz":
            ws_target_kz.append(row)
        elif pred == "ru":
            ws_target_ru.append(row)
        else:
            raise ValueError(f"Incorrect prediction: {pred}")

    # Save results
    target_file_kz = "target_kz.xlsx"
    target_file_ru = "target_ru.xlsx"
    wb_target_kz.save(target_file_kz)
    wb_target_ru.save(target_file_ru)

    print(f"Files saved: {target_file_kz} and {target_file_ru}")

    # Preview first few rows from each file if requested
    if preview_rows > 0:
        print("\nPreview of saved files:")

        print(f"\nFirst {preview_rows} rows from {target_file_kz}:")
        preview_file(target_file_kz, preview_rows)

        print(f"\nFirst {preview_rows} rows from {target_file_ru}:")
        preview_file(target_file_ru, preview_rows)


def preview_file(file_path: str, num_rows: int) -> None:
    """Preview first few rows of an Excel file.

    Args:
        file_path: Path to the Excel file
        num_rows: Number of rows to preview (including header)
    """
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        print(row)
        if i >= num_rows:
            break


if __name__ == "__main__":
    for url in urls:
        filename = url.split("/")[-1]
        response = requests.get(url)
        with open(filename, "wb") as f:
            f.write(response.content)
        print(f"Файл {filename} успешно скачан")

    # Чтение данных с обработкой NaN
    df_ru_names_1 = pd.read_excel("ru_names_1.xlsx", usecols=[0, 1, 2], header=None)
    df_kz_names_1 = pd.read_excel("kz_names_1.xlsx", usecols=[0, 1, 2], header=None)
    df_kz_names_2 = pd.read_excel("kz_names_2.xlsx", usecols=[3], header=None)
    # Преобразование DataFrame в списки
    list_ru_names_1 = df_to_string_list(df_ru_names_1)
    list_kz_names_1 = df_to_string_list(df_kz_names_1)
    list_kz_names_2 = df_to_string_list(
        df_kz_names_2
    )  # Уже содержит только один столбец

    # Объединение всех казахских имен в один список
    all_kz_names = list_kz_names_1 + list_kz_names_2

    # Проверка результатов
    print(f"Русских имен: {len(list_ru_names_1)}")
    print(f"Казахских имен: {len(all_kz_names)}")
    print("\nПримеры русских имен:", list_ru_names_1[:5])
    print("\nПримеры казахских имен:", all_kz_names[:5])

    data_kz = all_kz_names
    data_ru = list_ru_names_1

    # Создаем DataFrame
    df_kz = pd.DataFrame({"name": data_kz, "label": "kz"})
    df_ru = pd.DataFrame({"name": data_ru, "label": "ru"})
    df = pd.concat([df_kz, df_ru])

    # Векторизация + модель
    model = Pipeline(
        [
            (
                "tfidf",
                TfidfVectorizer(analyzer="char", ngram_range=(1, 3)),
            ),  # Анализируем символы и их сочетания
            ("clf", LogisticRegression()),
        ]
    )

    # Обучение
    X_train, X_test, y_train, y_test = train_test_split(
        df["name"], df["label"], test_size=0.2
    )
    model.fit(X_train, y_train)

    # Оценка
    print(f"Accuracy: {model.score(X_test, y_test):.2f}")

    # Предсказание для нового имени
    new_name = ["Олег Владимирович"]
    prediction = model.predict(new_name)
    print(f"{new_name[0]} -> {prediction[0]}")

    # Сохраняем модель в файл
    dump(model, "name_classifier.joblib")
